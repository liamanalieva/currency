import requests
import datetime
from openpyxl import load_workbook, Workbook
import os.path

if not os.path.isfile('data/currency_exchange.xlsx'):
    headers = ['Валюта', 'Курс', 'Единиц', 'Дата']
    workbook_name = 'data/currency_exchange.xlsx'
    wb = Workbook()
    page = wb.active
    page.title = 'Курсы валют'
    page.append(headers)
    wb.save(filename=workbook_name)

workbook_name = 'currency_exchange.xlsx'
wb = load_workbook(workbook_name)
page = wb.active

try:
    cb_response = requests.get('https://www.cbr-xml-daily.ru/daily_json.js')
except:
    print("Ошибка подключения!")

needed_currencies = ['USD', 'EUR', 'BYN', 'KZT']  # create a list with needed currencies

if cb_response.status_code == 200:
    cb_data = cb_response.json()

    currencies_list = cb_data["Valute"]

    for currency in currencies_list:
        if currency in needed_currencies:
            value = currencies_list[currency]["Value"]
            nominal = currencies_list[currency]["Nominal"]
            real_value = round(value / nominal, 15)
            name = currencies_list[currency]["Name"]
            date = datetime.date.today().strftime("%m.%d.%Y")
            row = [name, real_value, nominal, date]  # ряд

            page.append(row)

    wb.save(filename=workbook_name)
else:
    print("Ошибка сервера данных!")

